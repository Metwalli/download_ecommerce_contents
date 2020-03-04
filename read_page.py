import requests
from lxml import html
import urllib
import openpyxl as excel

def remove_fobidden_chr(txt):
    txt = txt.replace('/', '')
    txt = txt.replace('>', '')
    txt = txt.replace('<', '')
    txt = txt.replace(':', '')
    txt = txt.replace('"', '')
    txt = txt.replace('\\', '')
    txt = txt.replace('\n', '')
    txt = txt.replace('|', '')
    txt = txt.replace('?', '')
    txt = txt.replace('*', '')
    return txt

def shein(url, output_path):
    # the target we want to open


    # open with GET method
    resp = requests.get(url)


    # http_respone 200 means OK status
    if resp.status_code == 200:
        print("Successfully opened the web page")
        print("The news are as follow :-\n")
        tree = html.fromstring(resp.content.decode("utf-8"))

        # This will create a list of buyers:
        images = [i for i in tree.xpath('//img[contains(@class,"j-verlok-lazy")]')]
        titles = [t for t in tree.xpath('//a[contains(@class,"c-goodsitem__goods-name j-goodsitem__goods-name ")]')]
        prices = [p for p in tree.xpath('//span[contains(@text(),"S.R.")]')]
        # num_likes = [n for n in tree.xpath('//span[contains(@class,"J-dataNum")]')]
        for p in prices:
            print(p.get('text'))
        # This will create a list of prices
        # prices = tree.xpath('//p[@class="price"]/text()')
        for t, i, p in zip(titles, images, prices):
           image_link = i.get('data-original')
           title = remove_fobidden_chr(t.text)
           urllib.request.urlretrieve(image_link, output_path + p.text + "--" + title + ".jpg")
    else:
        print("Error")

def hibobi(url, output_path, idx):
    # the target we want to open


    # open with GET method
    resp = requests.get(url)


    # http_respone 200 means OK status
    if resp.status_code == 200:
        print("Successfully opened the web page")
        print("The news are as follow :-\n")
        tree = html.fromstring(resp.content.decode("utf-8"))

        # This will create a list of buyers:
        images = [i for i in tree.xpath('//a/span[contains(@class,"img_box_bg")]/img')]
        titles = [t for t in tree.xpath('//a/p[contains(@class, "producttitle")]')]
        prices = [p for p in tree.xpath('//a/p[contains(@class, "price")]')]

        # This will create a list of prices
        # prices = tree.xpath('//p[@class="price"]/text()')

        for t, i, p in zip(titles, images, prices):
           image_link = i.get('src').replace('?x-oss-process=image/auto-orient,1/resize,m_lfit,w_400,limit_0/quality,q_90', '')
           title = remove_fobidden_chr(t.text)
           price = remove_fobidden_chr(p.text).replace(' ', '')
           ws['A' + str(idx)] = title
           ws['B' + str(idx)] = price
           idx = idx + 1
           print("image:", idx)
           try:
               out_image = output_path + "--" + price + "--" + title + ".jpg"
               r = requests.get(image_link)
               if r.status_code == 200:
                   with open(out_image, 'wb') as f:
                       f.write(r.content)
           except Exception:
                pass
    else:
        print("Error")

    return idx

def patpat(url, output_path):
    # the target we want to open


    # open with GET method
    resp = requests.get(url)


    # http_respone 200 means OK status
    if resp.status_code == 200:
        print("Successfully opened the web page")
        print("The news are as follow :-\n")
        tree = html.fromstring(resp.content.decode("utf-8"))

        # This will create a list of buyers:
        images = [i for i in tree.xpath('//img[contains(@class,"img")]')]
        titles = [t for t in tree.xpath('//div[contains(@class, "info")]')]
        prices = [p for p in tree.xpath('//div[contains(@class, "info")]/p')]

        # This will create a list of prices
        # prices = tree.xpath('//p[@class="price"]/text()')
        for t, i, p in zip(titles, images, prices):
           image_link = i.get('data-original')
           title = remove_fobidden_chr(t.text)
           urllib.request.urlretrieve(image_link, output_path + "--" + p.text + "--" + title + ".jpg")
    else:
        print("Error")

def jollychic(url, output_path):
    # the target we want to open


    # open with GET method
    resp = requests.get(url)


    # http_respone 200 means OK status
    if resp.status_code == 200:
        print("Successfully opened the web page")
        print("The news are as follow :-\n")
        tree = html.fromstring(resp.content.decode("utf-8"))

        # This will create a list of buyers:
        images = [i for i in tree.xpath('//a/img[contains(@class,"J-lazy-load firstImg")]')]
        titles = [t for t in tree.xpath('//a/h4[contains(@class,"pro_list_msg_1")]')]
        prices = [p for p in tree.xpath('//div[contains(@class,"pro_list_price_1 categoryTwo-loveBox")]/b')]
        num_likes = [p for p in tree.xpath('//span[contains(@class,"J-dataNum")]')]

        # This will create a list of prices
        # prices = tree.xpath('//p[@class="price"]/text()')
        for t, i, p, n in zip(titles, images, prices, num_likes):
           image_link = i.get('data-original')
           title = remove_fobidden_chr(t.text)
           urllib.request.urlretrieve(image_link, output_path + str(n.text) + "--" + p.text + "--" + title + ".jpg")
    else:
        print("Error")
url = 'https://www.hibobi.com/categories/Toddler-Girl-Cate-83-relate-PAGE-2.html?spm=1001.2001.84-relate.3&lang=en&page=2'
output_path = 'images/hibobi/matching_outfits/'
pages = 4

wb = excel.Workbook()
ws = wb.active
ws['A1'] = "Item Name"
ws['B1'] = "Price"
idx = 2
for i in range(1, pages):
    url = 'https://www.hibobi.com/categories/Matching-Outfits-Cate-180-relate-PAGE-' + str(
        i) + '.html?spm=1001.2001.85-relate.6&lang=en&page=' + str(i)
    idx = hibobi(url, output_path, idx)
    print('Finish Page:', i)



wb.save("images/hibobi/matching_outfits.xlsx")